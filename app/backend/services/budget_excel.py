from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.backend.core.constants import (
    TRANSACTION_TYPE_EXPENSE,
    TRANSACTION_TYPE_INCOME,
    TRANSACTION_TYPE_TRANSFER,
)
from app.backend.core.security import decrypt_amount
from app.backend.models.budget import BudgetAccount, BudgetCategory, BudgetTransaction
from app.backend.models.user import User

MONTHS_RU = (
    "",
    "Январь",
    "Февраль",
    "Март",
    "Апрель",
    "Май",
    "Июнь",
    "Июль",
    "Август",
    "Сентябрь",
    "Октябрь",
    "Ноябрь",
    "Декабрь",
)

COL_SUMMARY = 1
COL_INCOME = 6
COL_EXPENSE = 11
INCOME_SUM_COL = COL_INCOME + 3
EXPENSE_SUM_COL = COL_EXPENSE + 3

FILL_SUMMARY_HEADER = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_SUMMARY_ALT = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
FILL_INCOME_HEADER = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
FILL_EXPENSE_HEADER = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")
FILL_INCOME_AMOUNT = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_EXPENSE_AMOUNT = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)


def _tx_amount(tx: BudgetTransaction) -> Decimal:
    if tx.amount_encrypted:
        return decrypt_amount(tx.amount_encrypted)
    return Decimal(str(tx.amount or 0))


def _tx_date(tx: BudgetTransaction) -> date:
    occurred = tx.occurred_at
    if hasattr(occurred, "date"):
        return occurred.date()
    return occurred


def _period_title(d1: date, d2: date) -> str:
    if d1.year == d2.year and d1.month == d2.month and d1.day == 1:
        last_day = (d2.replace(day=28) + timedelta(days=4)).replace(day=1)
        month_end = last_day - timedelta(days=1)
        if d2 == month_end:
            return MONTHS_RU[d1.month]
    return f"{d1.strftime('%d.%m.%Y')} — {d2.strftime('%d.%m.%Y')}"


def _fmt_rub_compact(value: Decimal | float) -> str:
    amount = float(value)
    sign = "-" if amount < 0 else ""
    whole = f"{abs(int(round(amount))):,}".replace(",", " ")
    return f"р.{sign}{whole}"


def _write_money_cell(ws, row: int, col: int, value: Decimal, *, fill: PatternFill | None = None, bold: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=float(value))
    cell.number_format = '#,##0" ₽"'
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if fill:
        cell.fill = fill
    if bold:
        cell.font = Font(bold=True)
    cell.border = THIN_BORDER


def _table_header(ws, row: int, col_start: int, headers: tuple[str, ...], fill: PatternFill, *, color: str = "FFFFFF") -> None:
    for i, title in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=title)
        c.fill = fill
        c.font = Font(bold=True, color=color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER


def _split_balance_totals(accounts: list[BudgetAccount], balances: dict[int, Decimal]) -> tuple[Decimal, Decimal]:
    regular = Decimal("0")
    savings = Decimal("0")
    for acc in accounts:
        value = balances.get(acc.id, Decimal("0"))
        if acc.is_savings:
            savings += value
        else:
            regular += value
    return regular, savings


def _apply_tx_to_balances(balances: dict[int, Decimal], tx: BudgetTransaction, amount: Decimal) -> None:
    if tx.type == TRANSACTION_TYPE_INCOME:
        if tx.account_id:
            balances[tx.account_id] += amount
    elif tx.type == TRANSACTION_TYPE_EXPENSE:
        if tx.account_id:
            balances[tx.account_id] -= amount
    elif tx.type == TRANSACTION_TYPE_TRANSFER:
        if tx.account_id:
            balances[tx.account_id] -= amount
        if tx.contra_account_id:
            balances[tx.contra_account_id] += amount


def _sheet_main(
    wb: Workbook,
    *,
    user: User,
    d1: date,
    d2: date,
    accounts: list[BudgetAccount],
    account_by_id: dict[int, BudgetAccount],
    category_map: dict[int, BudgetCategory],
    period_txs: list[BudgetTransaction],
    history_txs: list[BudgetTransaction],
) -> None:
    ws = wb.active
    ws.title = "Бюджет"

    start_balances: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
    end_balances: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
    for tx in history_txs:
        amount = _tx_amount(tx)
        tx_day = _tx_date(tx)
        _apply_tx_to_balances(end_balances, tx, amount)
        if tx_day < d1:
            _apply_tx_to_balances(start_balances, tx, amount)

    start_regular, start_savings = _split_balance_totals(accounts, start_balances)
    end_regular, end_savings = _split_balance_totals(accounts, end_balances)

    income_txs = [tx for tx in period_txs if tx.type == TRANSACTION_TYPE_INCOME]
    expense_txs = [tx for tx in period_txs if tx.type == TRANSACTION_TYPE_EXPENSE]

    income_total = sum((_tx_amount(tx) for tx in income_txs), Decimal("0"))
    expense_total = sum((_tx_amount(tx) for tx in expense_txs), Decimal("0"))
    net_total = income_total - expense_total

    savings_delta = Decimal("0")
    for tx in period_txs:
        amount = _tx_amount(tx)
        if tx.type == TRANSACTION_TYPE_INCOME:
            acc = account_by_id.get(tx.account_id or 0)
            if acc and acc.is_savings:
                savings_delta += amount
        elif tx.type == TRANSACTION_TYPE_EXPENSE:
            acc = account_by_id.get(tx.account_id or 0)
            if acc and acc.is_savings:
                savings_delta -= amount
        elif tx.type == TRANSACTION_TYPE_TRANSFER:
            acc_to = account_by_id.get(tx.contra_account_id or 0)
            acc_from = account_by_id.get(tx.account_id or 0)
            if acc_to and acc_to.is_savings:
                savings_delta += amount
            elif acc_from and acc_from.is_savings:
                savings_delta -= amount

    period_days = (d2 - d1).days + 1
    active_days = len({_tx_date(tx) for tx in period_txs})
    tx_count = len(period_txs)
    avg_income = income_total / Decimal(period_days) if period_days else Decimal("0")
    avg_expense = expense_total / Decimal(period_days) if period_days else Decimal("0")
    savings_rate = (income_total - expense_total) / income_total * Decimal("100") if income_total > 0 else Decimal("0")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 2
    for col in (COL_INCOME, COL_INCOME + 1, COL_INCOME + 2, INCOME_SUM_COL):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["J"].width = 2
    for col in (COL_EXPENSE, COL_EXPENSE + 1, COL_EXPENSE + 2, EXPENSE_SUM_COL):
        ws.column_dimensions[get_column_letter(col)].width = 16

    title = _period_title(d1, d2)
    ws.merge_cells(start_row=1, start_column=COL_SUMMARY, end_row=1, end_column=COL_SUMMARY + 3)
    title_cell = ws.cell(row=1, column=COL_SUMMARY, value=title)
    title_cell.font = Font(size=20, bold=True, color="1F4E78")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=2, start_column=COL_SUMMARY, end_row=2, end_column=COL_SUMMARY + 3)
    subtitle = ws.cell(row=2, column=COL_SUMMARY, value=f"{d1.strftime('%d.%m.%Y')} — {d2.strftime('%d.%m.%Y')}")
    subtitle.font = Font(size=10, color="6B7280")
    subtitle.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=3, start_column=COL_SUMMARY, end_row=3, end_column=COL_SUMMARY + 3)
    user_label = user.full_name or user.tg_username or user.email
    generated = ws.cell(
        row=3,
        column=COL_SUMMARY,
        value=f"Пользователь: {user_label} | Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
    )
    generated.font = Font(size=9, color="6B7280")
    generated.alignment = Alignment(horizontal="left", vertical="center")

    summary_header_row = 5
    _table_header(ws, summary_header_row, COL_SUMMARY, ("", "Счета", "Сбережения", "Итого"), FILL_SUMMARY_HEADER, color="000000")

    summary_rows = [
        ("Остаток на начало периода", start_regular, start_savings),
        ("Остаток на конец периода", end_regular, end_savings),
        ("Сбережения (изменение)", None, savings_delta),
        ("Доходов за период", income_total, None),
        ("Расходов за период", expense_total, None),
        ("Разница", net_total, None),
    ]

    for offset, row_data in enumerate(summary_rows, start=1):
        row = summary_header_row + offset
        label, col_b, col_c = row_data
        label_cell = ws.cell(row=row, column=COL_SUMMARY, value=label)
        label_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        label_cell.border = THIN_BORDER
        if offset % 2 == 0:
            label_cell.fill = FILL_SUMMARY_ALT

        if col_b is not None and col_c is not None:
            _write_money_cell(ws, row, COL_SUMMARY + 1, col_b)
            _write_money_cell(ws, row, COL_SUMMARY + 2, col_c)
            _write_money_cell(ws, row, COL_SUMMARY + 3, col_b + col_c, bold=label == "Разница")
        elif col_b is not None:
            ws.cell(row=row, column=COL_SUMMARY + 1).border = THIN_BORDER
            ws.cell(row=row, column=COL_SUMMARY + 2).border = THIN_BORDER
            _write_money_cell(ws, row, COL_SUMMARY + 3, col_b, bold=label == "Разница")
        else:
            ws.cell(row=row, column=COL_SUMMARY + 1).border = THIN_BORDER
            _write_money_cell(ws, row, COL_SUMMARY + 2, col_c)
            ws.cell(row=row, column=COL_SUMMARY + 3).border = THIN_BORDER

    kpi_header_row = summary_header_row + len(summary_rows) + 2
    _table_header(ws, kpi_header_row, COL_SUMMARY, ("KPI", "Значение", "", ""), FILL_SUMMARY_HEADER, color="000000")
    ws.merge_cells(start_row=kpi_header_row, start_column=COL_SUMMARY + 1, end_row=kpi_header_row, end_column=COL_SUMMARY + 3)

    kpis = [
        ("Транзакций", Decimal(tx_count), "integer"),
        ("Активных дней", Decimal(active_days), "integer"),
        ("Средний доход в день", avg_income, "money"),
        ("Средний расход в день", avg_expense, "money"),
        ("Норма сбережений", savings_rate, "percent"),
    ]

    for i, (label, value, kind) in enumerate(kpis, start=1):
        row = kpi_header_row + i
        cell = ws.cell(row=row, column=COL_SUMMARY, value=label)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=COL_SUMMARY + 1, end_row=row, end_column=COL_SUMMARY + 3)
        v = ws.cell(row=row, column=COL_SUMMARY + 1, value=float(value))
        v.border = THIN_BORDER
        v.alignment = Alignment(horizontal="right", vertical="center")
        if kind == "money":
            v.number_format = '#,##0" ₽"'
        elif kind == "percent":
            v.number_format = '0.0"%"'
        else:
            v.number_format = "0"

    account_header_row = kpi_header_row + len(kpis) + 2
    _table_header(ws, account_header_row, COL_SUMMARY, ("Счет", "Тип", "Остаток", "Доля"), FILL_SUMMARY_HEADER, color="000000")

    total_end = end_regular + end_savings
    row = account_header_row + 1
    for acc in accounts:
        balance = end_balances.get(acc.id, Decimal("0"))
        ws.cell(row=row, column=COL_SUMMARY, value=acc.title).border = THIN_BORDER
        ws.cell(row=row, column=COL_SUMMARY + 1, value="Сбережения" if acc.is_savings else "Счет").border = THIN_BORDER
        _write_money_cell(ws, row, COL_SUMMARY + 2, balance)

        share_cell = ws.cell(row=row, column=COL_SUMMARY + 3, value=float((balance / total_end * Decimal("100")) if total_end != 0 else 0))
        share_cell.number_format = '0.0"%"'
        share_cell.alignment = Alignment(horizontal="right", vertical="center")
        share_cell.border = THIN_BORDER
        row += 1

    section_row = 5
    data_row = section_row + 2

    for col_start, title_text, total, header_fill, columns in (
        (
            COL_INCOME,
            "Доходы",
            income_total,
            FILL_INCOME_HEADER,
            ("Дата", "Источник", "Категория", "Сумма"),
        ),
        (
            COL_EXPENSE,
            "Расходы",
            expense_total,
            FILL_EXPENSE_HEADER,
            ("Дата", "Назначение", "Категория", "Сумма"),
        ),
    ):
        for col in range(col_start, col_start + 3):
            cell = ws.cell(row=section_row, column=col)
            cell.fill = header_fill
            cell.border = THIN_BORDER

        title_cell = ws.cell(row=section_row, column=col_start, value=title_text)
        title_cell.fill = header_fill
        title_cell.font = Font(bold=True, color="FFFFFF", size=12)
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        title_cell.border = THIN_BORDER

        total_cell = ws.cell(row=section_row, column=col_start + 3, value=f"Итого: {_fmt_rub_compact(total)}")
        total_cell.fill = header_fill
        total_cell.font = Font(bold=True, color="FFFFFF", size=11)
        total_cell.alignment = Alignment(horizontal="right", vertical="center")
        total_cell.border = THIN_BORDER

        _table_header(ws, section_row + 1, col_start, columns, header_fill)

    income_row = data_row
    for tx in income_txs:
        category_name = category_map[tx.category_id].name if tx.category_id and tx.category_id in category_map else "—"
        ws.cell(row=income_row, column=COL_INCOME, value=_tx_date(tx).strftime("%d.%m.%Y")).border = THIN_BORDER
        ws.cell(row=income_row, column=COL_INCOME + 1, value=tx.description or "—").border = THIN_BORDER
        ws.cell(row=income_row, column=COL_INCOME + 2, value=category_name).border = THIN_BORDER
        _write_money_cell(ws, income_row, INCOME_SUM_COL, _tx_amount(tx), fill=FILL_INCOME_AMOUNT)
        income_row += 1

    expense_row = data_row
    for tx in expense_txs:
        category_name = category_map[tx.category_id].name if tx.category_id and tx.category_id in category_map else "—"
        ws.cell(row=expense_row, column=COL_EXPENSE, value=_tx_date(tx).strftime("%d.%m.%Y")).border = THIN_BORDER
        ws.cell(row=expense_row, column=COL_EXPENSE + 1, value=tx.description or "—").border = THIN_BORDER
        ws.cell(row=expense_row, column=COL_EXPENSE + 2, value=category_name).border = THIN_BORDER
        _write_money_cell(ws, expense_row, EXPENSE_SUM_COL, _tx_amount(tx), fill=FILL_EXPENSE_AMOUNT)
        expense_row += 1

    if income_row > data_row:
        income_range = f"{get_column_letter(INCOME_SUM_COL)}{data_row}:{get_column_letter(INCOME_SUM_COL)}{income_row - 1}"
        ws.conditional_formatting.add(
            income_range,
            ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="63BE7B"),
        )

    if expense_row > data_row:
        expense_range = f"{get_column_letter(EXPENSE_SUM_COL)}{data_row}:{get_column_letter(EXPENSE_SUM_COL)}{expense_row - 1}"
        ws.conditional_formatting.add(
            expense_range,
            ColorScaleRule(
                start_type="min",
                start_color="63BE7B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="F8696B",
            ),
        )

    if income_row == data_row:
        ws.merge_cells(start_row=data_row, start_column=COL_INCOME, end_row=data_row, end_column=INCOME_SUM_COL)
        no_income = ws.cell(row=data_row, column=COL_INCOME, value="Нет доходов за период")
        no_income.alignment = Alignment(horizontal="center", vertical="center")
        no_income.border = THIN_BORDER

    if expense_row == data_row:
        ws.merge_cells(start_row=data_row, start_column=COL_EXPENSE, end_row=data_row, end_column=EXPENSE_SUM_COL)
        no_expense = ws.cell(row=data_row, column=COL_EXPENSE, value="Нет расходов за период")
        no_expense.alignment = Alignment(horizontal="center", vertical="center")
        no_expense.border = THIN_BORDER

    ws.freeze_panes = f"{get_column_letter(COL_INCOME)}{data_row}"
    ws.sheet_view.showGridLines = True


def _sheet_operations(
    wb: Workbook,
    *,
    account_by_id: dict[int, BudgetAccount],
    category_map: dict[int, BudgetCategory],
    period_txs: list[BudgetTransaction],
) -> None:
    ws = wb.create_sheet("Операции")
    headers = ("Дата", "Тип", "Сумма", "Счет", "Контр. счет", "Категория", "Описание")
    _table_header(ws, 1, 1, headers, FILL_SUMMARY_HEADER, color="000000")

    type_labels = {
        TRANSACTION_TYPE_INCOME: "Доход",
        TRANSACTION_TYPE_EXPENSE: "Расход",
        TRANSACTION_TYPE_TRANSFER: "Перевод",
    }

    row = 2
    for tx in period_txs:
        ws.cell(row=row, column=1, value=_tx_date(tx).strftime("%d.%m.%Y")).border = THIN_BORDER
        ws.cell(row=row, column=2, value=type_labels.get(tx.type, tx.type)).border = THIN_BORDER

        sum_cell = ws.cell(row=row, column=3, value=float(_tx_amount(tx)))
        sum_cell.number_format = '#,##0" ₽"'
        sum_cell.alignment = Alignment(horizontal="right", vertical="center")
        sum_cell.border = THIN_BORDER

        ws.cell(row=row, column=4, value=(account_by_id.get(tx.account_id or 0).title if tx.account_id and tx.account_id in account_by_id else "—")).border = THIN_BORDER
        ws.cell(row=row, column=5, value=(account_by_id.get(tx.contra_account_id or 0).title if tx.contra_account_id and tx.contra_account_id in account_by_id else "—")).border = THIN_BORDER
        ws.cell(row=row, column=6, value=(category_map.get(tx.category_id or 0).name if tx.category_id and tx.category_id in category_map else "—")).border = THIN_BORDER
        ws.cell(row=row, column=7, value=tx.description or "—").border = THIN_BORDER
        row += 1

    for col_idx, width in {1: 13, 2: 10, 3: 14, 4: 22, 5: 22, 6: 24, 7: 44}.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"


def _sheet_analytics(
    wb: Workbook,
    *,
    d1: date,
    d2: date,
    category_map: dict[int, BudgetCategory],
    period_txs: list[BudgetTransaction],
) -> None:
    ws = wb.create_sheet("Аналитика")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 4
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 16

    ws.merge_cells("A1:E1")
    title = ws["A1"]
    title.value = f"Аналитика периода: {d1.strftime('%d.%m.%Y')} — {d2.strftime('%d.%m.%Y')}"
    title.font = Font(size=14, bold=True, color="1F4E78")
    title.alignment = Alignment(horizontal="left")

    income_by_cat: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    expense_by_cat: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    by_day_income: dict[date, Decimal] = defaultdict(lambda: Decimal("0"))
    by_day_expense: dict[date, Decimal] = defaultdict(lambda: Decimal("0"))

    for tx in period_txs:
        amount = _tx_amount(tx)
        tx_day = _tx_date(tx)
        cat_name = category_map.get(tx.category_id or 0).name if tx.category_id and tx.category_id in category_map else "Без категории"

        if tx.type == TRANSACTION_TYPE_INCOME:
            income_by_cat[cat_name] += amount
            by_day_income[tx_day] += amount
        elif tx.type == TRANSACTION_TYPE_EXPENSE:
            expense_by_cat[cat_name] += amount
            by_day_expense[tx_day] += amount

    _table_header(ws, 3, 1, ("Топ источники дохода", "Сумма"), FILL_INCOME_HEADER)
    _table_header(ws, 3, 4, ("Топ категории расхода", "Сумма"), FILL_EXPENSE_HEADER)

    top_income = sorted(income_by_cat.items(), key=lambda x: x[1], reverse=True)[:10]
    top_expense = sorted(expense_by_cat.items(), key=lambda x: x[1], reverse=True)[:10]
    max_rows = max(len(top_income), len(top_expense), 1)

    for i in range(max_rows):
        r = 4 + i
        if i < len(top_income):
            ws.cell(row=r, column=1, value=top_income[i][0]).border = THIN_BORDER
            _write_money_cell(ws, r, 2, top_income[i][1], fill=FILL_INCOME_AMOUNT)
        else:
            ws.cell(row=r, column=1, value="—").border = THIN_BORDER
            ws.cell(row=r, column=2).border = THIN_BORDER

        if i < len(top_expense):
            ws.cell(row=r, column=4, value=top_expense[i][0]).border = THIN_BORDER
            _write_money_cell(ws, r, 5, top_expense[i][1], fill=FILL_EXPENSE_AMOUNT)
        else:
            ws.cell(row=r, column=4, value="—").border = THIN_BORDER
            ws.cell(row=r, column=5).border = THIN_BORDER

    day_start = 6 + max_rows
    _table_header(ws, day_start, 7, ("Дата", "Доход", "Расход", "Дельта"), FILL_SUMMARY_HEADER, color="000000")

    row = day_start + 1
    day = d1
    while day <= d2:
        income = by_day_income.get(day, Decimal("0"))
        expense = by_day_expense.get(day, Decimal("0"))
        delta = income - expense

        ws.cell(row=row, column=7, value=day.strftime("%d.%m.%Y")).border = THIN_BORDER
        _write_money_cell(ws, row, 8, income, fill=FILL_INCOME_AMOUNT)
        _write_money_cell(ws, row, 9, expense, fill=FILL_EXPENSE_AMOUNT)
        _write_money_cell(ws, row, 10, delta)
        row += 1
        day += timedelta(days=1)

    ws.freeze_panes = "A4"


def build_budget_excel_bytes(
    db: Session,
    user: User,
    d1: date,
    d2: date,
) -> bytes:
    wb = Workbook()

    accounts = list(
        db.scalars(select(BudgetAccount).where(BudgetAccount.user_id == user.id).order_by(BudgetAccount.id.asc())).all()
    )
    account_by_id = {acc.id: acc for acc in accounts}
    category_map = {
        c.id: c for c in db.scalars(select(BudgetCategory).where(BudgetCategory.user_id == user.id)).all()
    }

    period_txs = list(
        db.scalars(
            select(BudgetTransaction)
            .where(
                BudgetTransaction.user_id == user.id,
                BudgetTransaction.occurred_at >= d1,
                BudgetTransaction.occurred_at <= d2,
            )
            .order_by(BudgetTransaction.occurred_at.asc(), BudgetTransaction.id.asc())
        ).all()
    )

    history_txs = list(
        db.scalars(
            select(BudgetTransaction)
            .where(
                BudgetTransaction.user_id == user.id,
                BudgetTransaction.occurred_at <= d2,
            )
            .order_by(BudgetTransaction.occurred_at.asc(), BudgetTransaction.id.asc())
        ).all()
    )

    _sheet_main(
        wb,
        user=user,
        d1=d1,
        d2=d2,
        accounts=accounts,
        account_by_id=account_by_id,
        category_map=category_map,
        period_txs=period_txs,
        history_txs=history_txs,
    )
    _sheet_analytics(wb, d1=d1, d2=d2, category_map=category_map, period_txs=period_txs)
    _sheet_operations(wb, account_by_id=account_by_id, category_map=category_map, period_txs=period_txs)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()
